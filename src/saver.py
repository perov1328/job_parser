
import csv
import openpyxl
import json
from src.abstractions import Saver
from json import JSONDecodeError
from src.vacancy import Vacancy


class JSONSaver(Saver):
    """
    Класс для сохранения результатов в json формате
    """
    def __init__(self, path, vacancies = None):
        self.path = path
        self.dict_vacancies = vacancies
        self.instance_vacancies = []

    def save_to_file(self):
        """
        Сохранение файла
        """
        with open(self.path, "w", encoding="utf-8") as file:
            json.dump(self.dict_vacancies, file, indent=4)

    def load_from_file(self):
        """
        Загрузка данных о вакансии из файла
        """
        try:

            try:
                with open(self.path, "r", encoding="utf-8") as file:
                    data = json.load(file)
            except JSONDecodeError:
                print("Файл пуст")
                return

            for el in data:
                vacancy = Vacancy(
                    int(el.get('id_num', "Нет данных")),
                    str(el.get('vacancy_name', "Нет данных")),
                    str(el.get('company_name')),
                    str(el.get('description', "Нет данных")),
                    int(el.get('salary_from')) if el.get('salary_from') is not None else None,
                    int(el.get('salary_to')) if el.get('salary_to') is not None else None,
                    str(el.get('salary_currency')),
                    str(el.get('area', "Нет данных")),
                    str(el.get('url', "Нет данных")),
                )
                self.instance_vacancies.append(vacancy)
            self.dict_vacancies = [vacancy.to_dict() for vacancy in self.instance_vacancies]
        except FileNotFoundError:
            self.dict_vacancies = []
            raise FileNotFoundError("File not found")

    def add_vacancy(self, vacancy):
        """
        Добавление вакансии
        """
        self.load_from_file()
        self.dict_vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        """
        Поиск и получение вакансии
        """
        self.load_from_file()
        result = []
        for vacancy in self.instance_vacancies:
            dict_vacancy = vacancy.to_dict()
            if criteria in dict_vacancy['vacancy_name'] or criteria in dict_vacancy['description']:
                result.append(vacancy)
        return result

    def remove_vacancy(self, vacancy_id):
        """
        Удаление вакансии
        """
        self.load_from_file()
        try:
            self.dict_vacancies = [vacancy for vacancy in self.dict_vacancies if vacancy['id_num'] != int(vacancy_id)]
            self.save_to_file()
            print("Вакансия удалена!")
        except ValueError:
            print("Вакансии с таким номером не существует.")


class CSVSaver(Saver):
    """
    Класс для сохранения результатов в CSV формате
    """
    def __init__(self, filename: str, vacancies: list):
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        """
        Сохранение файла
        """
        with open(self.filename, 'w', newline='', encoding='utf-8') as file:
            fieldnames = ['id_num', 'vacancy_name', 'company_name', 'description',
                          'salary_from', 'salary_to', 'salary_currency',
                          'area', 'url']
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.vacancies)

    def load_from_file(self):
        """
        Загрузка данных о вакансии из файла
        """
        try:
            with open(self.filename, 'r', newline='', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                self.vacancies = [el for el in reader]
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        """
        Добавление вакансии
        """
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        """
        Поиск и получение вакансии
        """
        self.load_from_file()
        result = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                result.append(vacancy)
        return result

    def remove_vacancy(self, vacancy_id):
        """
        Удаление вакансии
        """
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()


class ExcelSaver(Saver):
    """
    Класс для сохранения результатов в xls формате
    """
    def __init__(self, filename: str, vacancies: list):
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        """
        Сохранение файла
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['id', 'vacancy_name', 'company_name', 'description',
                      'salary_from', 'salary_to', 'salary_currency',
                      'area', 'url'])

        for vacancy in self.vacancies:
            sheet.append([vacancy['id_num'], vacancy['vacancy_name'], vacancy['company_name'],
                          vacancy['description'], vacancy['salary_from'], vacancy['salary_to'],
                          vacancy['salary_currency'], vacancy['area'], vacancy['url']])

        workbook.save(self.filename)

    def load_from_file(self):
        """
        Загрузка данных о вакансии из файла
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active
            self.vacancies = [{'id_num': row[0].value, 'vacancy_name': row[1].value, 'company_name': row[2].value,
                               'description': row[3].value, 'salary_from': row[4].value, 'salary_to': row[5].value,
                               'salary_currency': row[6].value, 'area': row[7].value, 'url': row[8].value}
                              for row in sheet.iter_rows(min_row=2)]
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        """
        Добавление вакансии
        """
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        """
        Поиск и получение вакансии
        """
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        """
        Удаление вакансии
        """
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()

class TxtSaver(Saver):
    """
    Класс для сохранения результатов в txt формате
    """
    def __init__(self, filename: str, vacancies: list):
        self.filename = filename
        self.vacancies = vacancies

    def save_to_file(self):
        """
        Сохранение файла
        """
        with open(self.filename, 'w', encoding='utf-8') as file:
            for vacancy in self.vacancies:
                file.write(f"Id: {vacancy['id_num']}\n")
                file.write(f"Vacancy Name: {vacancy['vacancy_name']}\n")
                file.write(f"Company Name: {vacancy['company_name']}\n")
                file.write(f"Description: {vacancy['description']}\n")
                file.write(f"Salary From: {vacancy['salary_from']}\n")
                file.write(f"Salary To: {vacancy['salary_to']}\n")
                file.write(f"Salary Currency: {vacancy['salary_currency']}\n")
                file.write(f"Area: {vacancy['area']}\n")
                file.write(f"URL: {vacancy['url']}\n\n")

    def load_from_file(self):
        """
        Загрузка данных о вакансии из файла
        """
        try:
            with open(self.filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()
                vacancy = {}
                for line in lines:
                    if line.startswith("Id: "):
                        vacancy['id'] = line.strip()[4:]
                    elif line.startswith("Title: "):
                        vacancy['title'] = line.strip()[7:]
                    elif line.startswith("Description: "):
                        vacancy['description'] = line.strip()[12:]
                    elif line == "\n":
                        self.vacancies.append(vacancy)
                        vacancy = {}
        except FileNotFoundError:
            self.vacancies = []

    def add_vacancy(self, vacancy):
        """
        Добавление вакансии
        """
        self.load_from_file()
        self.vacancies.append(vacancy)
        self.save_to_file()

    def get_vacancies(self, criteria):
        """
        Поиск и получение вакансии
        """
        self.load_from_file()
        matching_vacancies = []
        for vacancy in self.vacancies:
            if criteria in vacancy['title'] or criteria in vacancy['description']:
                matching_vacancies.append(vacancy)
        return matching_vacancies

    def remove_vacancy(self, vacancy_id):
        """
        Удаление вакансии
        """
        self.load_from_file()
        self.vacancies = [vacancy for vacancy in self.vacancies if vacancy['id_num'] != vacancy_id]
        self.save_to_file()
